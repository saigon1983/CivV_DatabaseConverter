import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

DB_ORIGINAL_PATH = r'Original/Tables'
DB_MODIFIED_PATH = r'Modified/Tables'

def getTable(tablename, path):
	# Возвращает любой открытый excel-файл с указанным названием по указанному пути
	return openpyxl.load_workbook(r'{}/{}'.format(path, tablename))
def getTableByName(tablename, original = True):
	# Возвращает оригинальную или модифицированную таблицу по названию без префиксов и расширений файла
	dp_path = DB_ORIGINAL_PATH if original else DB_MODIFIED_PATH
	suffix  = '_ORG' if original else '_MOD'
	return openpyxl.load_workbook(r'{}/NewTU_{}{}.xlsx'.format(dp_path, tablename, suffix))
def getProperTableName(tableName, sheetName):
	# сопоставляет наименованию каждой вкладки excel-файла правильное наименование таблицы в базе данных
	prefixes = {'BuildingClasses': 		'BuildingClass',
				'Buildings': 			'Building',
				'Civilizations': 		'Civilization',
				'Eras': 				'Era',
				'Features': 			'Feature',
				'FakeFeatures': 		'Feature',
				'Improvements': 		'Improvement',
				'Leaders': 				'Leader',
				'Policies': 			'Policy',
				'PolicyBranchTypes':	'PolicyBranch',
				'Resources': 			'Resource',
				'Technologies': 		'Technology',
				'UnitClasses': 			'UnitClass',
				'Units': 				'Unit'}
	result = sheetName if sheetName in prefixes.keys() else '{}_{}'.format(prefixes[tableName], sheetName)
	if result == 'Building_DomainFreeExperiencePerGW': 		result = 'Building_DomainFreeExperiencePerGreatWork'
	if result == 'Policy_BuildinClassProductionModifiers': 	result = 'Policy_BuildingClassProductionModifiers'
	if result == 'Civilization__Leaders': 	result = 'Civilization_Leaders'
	return result
def getHeades(workSheet):
	# Возвращает словарь заголовков таблицы в виде <Номер ячейки> => <Заголовок столбца>
	result = {}
	for cell in workSheet[1]: result[cell.column] = cell.value
	return result
def parseCiv5Table(tablename, original = True):
	WB 			= getTableByName(tablename, original)	# Открытый файл excel-таблицы
	ENTITIES 	= {}									# Результирующий словарь сущностей

	SHEETS 		= {}									# Словарь вкладок <Название вкладки> => <Название таблицы>
	for sheetName in WB.sheetnames: SHEETS[sheetName] = getProperTableName(tablename, sheetName)

	for sheetName, tableName in SHEETS.items():
		WS 			= WB[sheetName]
		HEADERS 	= getHeades(WS)

		if sheetName == tableName:
			for row_index in range(3, WS.max_row + 1):
				entityLine = {}
				for col_index, header in HEADERS.items():
					resultLine = {}
					if header:
						address 	= '{}{}'.format(get_column_letter(col_index), row_index)
						entityLine[header] = WS[address].value
				entityName = entityLine['Type']
				if entityName not in ENTITIES.keys(): ENTITIES[entityName] = {}
				ENTITIES[entityName][tableName] = entityLine

		if sheetName != tableName:
			firstHeader 	= WS['A1'].value
			secondName 		= WS['A2'].value
			thirdName 		= WS['B1'].value if not WS['B2'].value else None

			for row_index in range(3, WS.max_row + 1):
				entityName	= WS[f'A{row_index}'].value
				wholeLine	= False
				resultLine 	= {}

				if entityName not in ENTITIES.keys(): 			  raise	ValueError
				if tableName  not in ENTITIES[entityName].keys(): ENTITIES[entityName][tableName] = []

				for col_index, header in HEADERS.items():
					col_letter 	= get_column_letter(col_index)
					address 	= f'{col_letter}{row_index}'
					cellValue 	= WS[address].value

					if secondName:
						if cellValue:
							if cellValue != entityName:
								result = {firstHeader: entityName, }
								if thirdName and col_letter == 'B': continue
								if thirdName: result[thirdName] = WS[f'B{row_index}'].value
								result[secondName] = WS[f'{col_letter}2'].value
								if cellValue != '+': result[header] = cellValue
								ENTITIES[entityName][tableName].append(result)
					else:
						if not wholeLine and col_index > 1 and cellValue: wholeLine = True
						if wholeLine and cellValue and col_index > 1: resultLine[header] = cellValue
				if resultLine:
					result = {firstHeader: entityName, }
					result.update(resultLine)
					ENTITIES[entityName][tableName].append(result)

		if sheetName == '_Leaders':
			for k, v in ENTITIES.items(): print(k, v)


	return ENTITIES